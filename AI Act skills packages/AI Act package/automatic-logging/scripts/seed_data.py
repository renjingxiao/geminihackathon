"""
Seed mock data for all 4 sheets and regenerate formatted Excel.
"""
import sys
from pathlib import Path

# Add scripts to path
sys.path.insert(0, str(Path(__file__).parent))

from utils import (
    ensure_output_dir, append_to_csv, consolidate_to_excel,
    CAPABILITY_CSV, BIOMETRIC_CSV, OPERATIONAL_CSV, RISK_CSV,
    CAPABILITY_HEADERS, BIOMETRIC_HEADERS
)
from logger import ComplianceLogger
from datetime import datetime, timedelta
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo


def seed_capability_checklist():
    """Add complete capability checklist data."""
    # Clear and reinitialize
    if CAPABILITY_CSV.exists():
        CAPABILITY_CSV.unlink()
    
    import csv
    with open(CAPABILITY_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(CAPABILITY_HEADERS)
        
        # Mock data matching Article 12 requirements
        data = [
            ["Automatic Recording Capability", 
             "System technically capable of automatically recording of events (logs).",
             "Compliant", 
             "ComplianceLogger class with @log_interaction decorator auto-captures all AI interactions."],
            ["Traceability of Functioning", 
             "Ensuring a level of traceability appropriate to the intended purpose.",
             "Compliant", 
             "SHA-256 hashing of inputs/outputs with unique UUIDs for each log entry."],
            ["Continuous Recording Lifecycle", 
             "Recording of events over the lifetime of the system.",
             "Compliant", 
             "Persistent CSV storage with Excel consolidation for long-term retention."],
            ["Risk Identification Capability", 
             "Ability to identify situations resulting in Article 79(1) risks.",
             "Compliant", 
             "AI Watchdog monitors latency (>5s) and DOS attacks (>10 calls/sec) in real-time."],
            ["Substantial Modification Detection", 
             "Ability to identify situations presenting substantial modifications.",
             "Compliant", 
             "System version tracking in every log entry enables modification detection."],
            ["Post-Market Monitoring Facilitation", 
             "Facilitating monitoring of operations (Art 26(5)) and post-market (Art 72).",
             "Compliant", 
             "AI Auditor generates compliance reports using Gemini LLM for pattern analysis."],
        ]
        writer.writerows(data)
    print(f"✓ Seeded {len(data)} rows to Capability Checklist")


def seed_biometric_specifics():
    """Add complete biometric specifics data."""
    if BIOMETRIC_CSV.exists():
        BIOMETRIC_CSV.unlink()
    
    import csv
    with open(BIOMETRIC_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["Requirement (Annex III 1(a))", "Log Field Name", "Data Format", "Retention Period (>6 Months)"])
        
        data = [
            ["Period of Use (Start/End)", "timestamp_start, timestamp_end", "DateTime (ISO 8601)", "6 Months"],
            ["Reference Database Checked", "system_version", "String/ID", "6 Months"],
            ["Input Data (Match Result)", "input_hash", "SHA-256 Hash (64 chars)", "6 Months"],
            ["Output Data Reference", "output_hash", "SHA-256 Hash (64 chars)", "6 Months"],
            ["Person Involved (Identification)", "log_id", "UUID v4", "6 Months"],
        ]
        writer.writerows(data)
    print(f"✓ Seeded {len(data)} rows to Biometric Specifics")


def seed_operational_logs():
    """Add sample operational logs."""
    logger = ComplianceLogger(system_version="1.0.0-demo")
    
    # Simulate various operations
    operations = [
        ("What is Article 12 of EU AI Act?", "Article 12 requires automatic logging..."),
        ("Analyze this document for compliance", "Document analysis complete: 3 issues found"),
        ("Generate compliance report", "Report generated successfully"),
        ("Check system health", "All systems operational"),
        ("Process user query about GDPR", "GDPR response provided"),
    ]
    
    for inp, out in operations:
        start = datetime.now()
        end = start + timedelta(seconds=0.5)
        logger.log_operation(inp, out, start, end)
    
    print(f"✓ Seeded {len(operations)} operational logs")


def seed_risk_logs():
    """Add sample risk logs."""
    logger = ComplianceLogger(system_version="1.0.0-demo")
    
    risks = [
        ("Risk", "Performance degradation / Availability Risk", "High latency detected: 6.2s", "Logged for review"),
        ("Incident", "Cybersecurity / Denial of Service", "10+ requests in 1 second", "Rate limiting applied"),
        ("Modification", "System Update", "Model updated from v1.0 to v1.1", "Version tracked"),
        ("Risk", "Data Quality", "Input validation failed for 3 records", "Records flagged"),
    ]
    
    for event_type, category, desc, action in risks:
        logger.log_risk(event_type, category, desc, action)
    
    print(f"✓ Seeded {len(risks)} risk logs")


def format_excel_tables():
    """Format Excel sheets as proper tables."""
    from utils import OUTPUT_EXCEL
    
    wb = openpyxl.load_workbook(OUTPUT_EXCEL)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        
        if max_row < 2:
            continue  # Skip empty sheets
        
        # Create table reference
        table_ref = f"A1:{openpyxl.utils.get_column_letter(max_col)}{max_row}"
        
        # Create safe table name (no spaces/special chars)
        table_name = sheet_name.replace(" ", "_").replace(".", "_").replace("&", "and")
        table_name = f"Table_{table_name}"
        
        # Create table
        table = Table(displayName=table_name, ref=table_ref)
        
        # Style
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        # Add table to worksheet
        ws.add_table(table)
        
        # Auto-fit columns (approximate)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    wb.save(OUTPUT_EXCEL)
    print(f"✓ Formatted {len(wb.sheetnames)} sheets as tables")


def main():
    print("=" * 50)
    print("Seeding Complete Mock Data")
    print("=" * 50)
    
    ensure_output_dir()
    
    # Seed all 4 sheets
    seed_capability_checklist()
    seed_biometric_specifics()
    seed_operational_logs()
    seed_risk_logs()
    
    # Consolidate to Excel
    print("\n" + "-" * 50)
    consolidate_to_excel()
    print("✓ Consolidated to Excel")
    
    # Format as tables
    format_excel_tables()
    
    print("\n" + "=" * 50)
    print("DONE! Check Output/Record_Keeping_Logging_Art12.xlsx")
    print("=" * 50)


if __name__ == "__main__":
    main()

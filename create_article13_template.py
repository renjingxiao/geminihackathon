#!/usr/bin/env python3
"""
Create EU AI Act Article 13 Compliance Template
Transparency and provision of information to deployers
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_article13_template():
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    subheader_font = Font(bold=True, size=11, color="FFFFFF")
    section_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    section_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    alt_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # ==================== Sheet 1: Instructions for Use ====================
    ws1 = wb.create_sheet("Instructions for Use")
    
    # Set column widths
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 35
    ws1.column_dimensions['C'].width = 50
    ws1.column_dimensions['D'].width = 40
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 30
    
    # Title
    ws1.merge_cells('A1:F1')
    ws1['A1'] = "EU AI ACT - ARTICLE 13: TRANSPARENCY AND PROVISION OF INFORMATION TO DEPLOYERS"
    ws1['A1'].font = header_font
    ws1['A1'].fill = header_fill
    ws1['A1'].alignment = center_alignment
    ws1.row_dimensions[1].height = 30
    
    # Subtitle
    ws1.merge_cells('A2:F2')
    ws1['A2'] = "Instructions for Use - Compliance Documentation Template"
    ws1['A2'].font = subheader_font
    ws1['A2'].fill = subheader_fill
    ws1['A2'].alignment = center_alignment
    
    # Headers
    headers = ["Ref.", "Requirement", "Description/Evidence", "Your Input", "Status", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Article 13 Requirements Data
    requirements = [
        # Section: Provider Information (Article 13(3)(a))
        ("Section", "PROVIDER INFORMATION - Article 13(3)(a)", "", "", "", ""),
        ("13.3.a.1", "Provider Name", "Legal name of the AI system provider", "", "☐ Complete", ""),
        ("13.3.a.2", "Provider Address", "Registered business address", "", "☐ Complete", ""),
        ("13.3.a.3", "Provider Contact Details", "Email, phone, and contact person", "", "☐ Complete", ""),
        ("13.3.a.4", "Authorised Representative (if applicable)", "Name and contact details of EU representative", "", "☐ Complete", "If provider is outside EU"),
        
        # Section: Intended Purpose (Article 13(3)(b)(i))
        ("Section", "INTENDED PURPOSE - Article 13(3)(b)(i)", "", "", "", ""),
        ("13.3.b.i.1", "Primary Intended Purpose", "Main use case and application of the AI system", "", "☐ Complete", ""),
        ("13.3.b.i.2", "Target Users/Deployers", "Who is intended to use this system", "", "☐ Complete", ""),
        ("13.3.b.i.3", "Target Subjects", "Who the AI system is intended to be used on", "", "☐ Complete", ""),
        ("13.3.b.i.4", "Deployment Context", "Environment and context of intended use", "", "☐ Complete", ""),
        ("13.3.b.i.5", "Precluded Uses", "Uses that are explicitly NOT intended", "", "☐ Complete", ""),
        
        # Section: Accuracy and Performance (Article 13(3)(b)(ii))
        ("Section", "ACCURACY, ROBUSTNESS & CYBERSECURITY - Article 13(3)(b)(ii)", "", "", "", ""),
        ("13.3.b.ii.1", "Accuracy Level", "Declared accuracy metrics and values", "", "☐ Complete", ""),
        ("13.3.b.ii.2", "Accuracy Metrics Used", "Specific metrics (e.g., precision, recall, F1)", "", "☐ Complete", ""),
        ("13.3.b.ii.3", "Robustness Level", "System resilience to errors/faults", "", "☐ Complete", ""),
        ("13.3.b.ii.4", "Cybersecurity Measures", "Security measures implemented", "", "☐ Complete", ""),
        ("13.3.b.ii.5", "Testing Conditions", "Conditions under which metrics were tested", "", "☐ Complete", ""),
        ("13.3.b.ii.6", "Circumstances Affecting Performance", "Known factors that may impact accuracy/robustness", "", "☐ Complete", ""),
        
        # Section: Known Risks (Article 13(3)(b)(iii))
        ("Section", "KNOWN RISKS - Article 13(3)(b)(iii)", "", "", "", ""),
        ("13.3.b.iii.1", "Health Risks", "Known circumstances that may cause health risks", "", "☐ Complete", ""),
        ("13.3.b.iii.2", "Safety Risks", "Known circumstances that may cause safety risks", "", "☐ Complete", ""),
        ("13.3.b.iii.3", "Fundamental Rights Risks", "Known circumstances affecting fundamental rights", "", "☐ Complete", ""),
        ("13.3.b.iii.4", "Foreseeable Misuse Scenarios", "Reasonably foreseeable misuse and associated risks", "", "☐ Complete", ""),
        
        # Section: Explainability (Article 13(3)(b)(iv))
        ("Section", "EXPLAINABILITY - Article 13(3)(b)(iv)", "", "", "", ""),
        ("13.3.b.iv.1", "Explainability Capabilities", "Technical means to explain AI outputs", "", "☐ Complete", "If applicable"),
        ("13.3.b.iv.2", "Interpretation Methods", "Methods available to interpret decisions", "", "☐ Complete", ""),
        
        # Section: Performance on Groups (Article 13(3)(b)(v))
        ("Section", "PERFORMANCE ON SPECIFIC GROUPS - Article 13(3)(b)(v)", "", "", "", ""),
        ("13.3.b.v.1", "Target Groups Identified", "Specific persons/groups the system is used on", "", "☐ Complete", "When appropriate"),
        ("13.3.b.v.2", "Group-Specific Performance", "Performance metrics per identified group", "", "☐ Complete", ""),
        
        # Section: Input Data Specifications (Article 13(3)(b)(vi))
        ("Section", "INPUT DATA SPECIFICATIONS - Article 13(3)(b)(vi)", "", "", "", ""),
        ("13.3.b.vi.1", "Required Input Data Format", "Technical specifications for input data", "", "☐ Complete", ""),
        ("13.3.b.vi.2", "Training Data Information", "Description of training datasets used", "", "☐ Complete", ""),
        ("13.3.b.vi.3", "Validation Data Information", "Description of validation datasets used", "", "☐ Complete", ""),
        ("13.3.b.vi.4", "Testing Data Information", "Description of testing datasets used", "", "☐ Complete", ""),
        
        # Section: Output Interpretation (Article 13(3)(b)(vii))
        ("Section", "OUTPUT INTERPRETATION - Article 13(3)(b)(vii)", "", "", "", ""),
        ("13.3.b.vii.1", "Output Format Description", "Format and structure of AI system outputs", "", "☐ Complete", ""),
        ("13.3.b.vii.2", "Interpretation Guidance", "How to correctly interpret outputs", "", "☐ Complete", ""),
        ("13.3.b.vii.3", "Appropriate Use of Outputs", "Guidance on proper use of AI outputs", "", "☐ Complete", ""),
        
        # Section: Pre-determined Changes (Article 13(3)(c))
        ("Section", "PRE-DETERMINED CHANGES - Article 13(3)(c)", "", "", "", ""),
        ("13.3.c.1", "Planned System Changes", "Changes pre-determined at conformity assessment", "", "☐ Complete", "If any"),
        ("13.3.c.2", "Performance Changes", "Expected performance variations", "", "☐ Complete", ""),
        
        # Section: Human Oversight (Article 13(3)(d))
        ("Section", "HUMAN OVERSIGHT MEASURES - Article 13(3)(d)", "", "", "", ""),
        ("13.3.d.1", "Human Oversight Requirements", "Required human oversight per Article 14", "", "☐ Complete", ""),
        ("13.3.d.2", "Technical Measures for Oversight", "Tools facilitating output interpretation", "", "☐ Complete", ""),
        ("13.3.d.3", "Override Mechanisms", "How to override or interrupt the system", "", "☐ Complete", ""),
        ("13.3.d.4", "Required Competencies", "Skills/training needed for oversight personnel", "", "☐ Complete", ""),
        
        # Section: Computational Resources (Article 13(3)(e))
        ("Section", "RESOURCES AND MAINTENANCE - Article 13(3)(e)", "", "", "", ""),
        ("13.3.e.1", "Computational Resources", "Required computing power/resources", "", "☐ Complete", ""),
        ("13.3.e.2", "Hardware Resources", "Required hardware specifications", "", "☐ Complete", ""),
        ("13.3.e.3", "Expected Lifetime", "Expected operational lifetime of system", "", "☐ Complete", ""),
        ("13.3.e.4", "Maintenance Measures", "Required maintenance activities", "", "☐ Complete", ""),
        ("13.3.e.5", "Maintenance Frequency", "How often maintenance should occur", "", "☐ Complete", ""),
        ("13.3.e.6", "Software Update Requirements", "Update procedures and frequency", "", "☐ Complete", ""),
        
        # Section: Logging (Article 13(3)(f))
        ("Section", "LOGGING MECHANISMS - Article 13(3)(f)", "", "", "", ""),
        ("13.3.f.1", "Log Collection Mechanism", "How logs are collected", "", "☐ Complete", ""),
        ("13.3.f.2", "Log Storage Requirements", "How and where logs should be stored", "", "☐ Complete", ""),
        ("13.3.f.3", "Log Interpretation Guidance", "How to interpret logged data", "", "☐ Complete", ""),
        ("13.3.f.4", "Log Retention Period", "How long logs must be retained", "", "☐ Complete", "Min. 6 months per Art. 19"),
    ]
    
    row = 5
    for req in requirements:
        if req[0] == "Section":
            ws1.merge_cells(f'A{row}:F{row}')
            ws1.cell(row=row, column=1, value=req[1])
            ws1.cell(row=row, column=1).font = section_font
            ws1.cell(row=row, column=1).fill = section_fill
            ws1.cell(row=row, column=1).alignment = Alignment(vertical='center')
            for col in range(1, 7):
                ws1.cell(row=row, column=col).border = thin_border
        else:
            for col, value in enumerate(req, 1):
                cell = ws1.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.alignment = wrap_alignment
                cell.border = thin_border
                if row % 2 == 0:
                    cell.fill = alt_fill
        row += 1
    
    # ==================== Sheet 2: Compliance Checklist ====================
    ws2 = wb.create_sheet("Compliance Checklist")
    
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 30
    
    ws2.merge_cells('A1:E1')
    ws2['A1'] = "ARTICLE 13 COMPLIANCE CHECKLIST"
    ws2['A1'].font = header_font
    ws2['A1'].fill = header_fill
    ws2['A1'].alignment = center_alignment
    
    checklist_headers = ["Article Ref.", "Requirement", "Compliant?", "Evidence Location", "Responsible Person"]
    for col, header in enumerate(checklist_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    checklist_items = [
        ("Art. 13(1)", "AI system designed for sufficient transparency", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(1)", "Deployers can interpret system outputs", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(1)", "Deployers can use outputs appropriately", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(2)", "Instructions for use provided in digital format", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(2)", "Information is concise", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(2)", "Information is complete", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(2)", "Information is correct", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(2)", "Information is clear", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(2)", "Information is relevant", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(2)", "Information is accessible", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(2)", "Information is comprehensible to deployers", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(3)(a)", "Provider identity included", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(3)(a)", "Provider contact details included", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(3)(b)(i)", "Intended purpose documented", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(3)(b)(ii)", "Accuracy levels declared", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(3)(b)(ii)", "Robustness information provided", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(3)(b)(ii)", "Cybersecurity information provided", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(3)(b)(iii)", "Known risk circumstances documented", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(3)(b)(iv)", "Explainability capabilities described", "☐ Yes / ☐ No / ☐ N/A", "", ""),
        ("Art. 13(3)(b)(v)", "Group-specific performance documented", "☐ Yes / ☐ No / ☐ N/A", "", ""),
        ("Art. 13(3)(b)(vi)", "Input data specifications provided", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(3)(b)(vii)", "Output interpretation guidance provided", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(3)(c)", "Pre-determined changes documented", "☐ Yes / ☐ No / ☐ N/A", "", ""),
        ("Art. 13(3)(d)", "Human oversight measures described", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(3)(e)", "Resource requirements specified", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(3)(e)", "Maintenance measures documented", "☐ Yes / ☐ No", "", ""),
        ("Art. 13(3)(f)", "Logging mechanisms described", "☐ Yes / ☐ No", "", ""),
    ]
    
    for row, item in enumerate(checklist_items, 4):
        for col, value in enumerate(item, 1):
            cell = ws2.cell(row=row, column=col, value=value)
            cell.font = normal_font
            cell.alignment = wrap_alignment
            cell.border = thin_border
            if row % 2 == 0:
                cell.fill = alt_fill
    
    # ==================== Sheet 3: Metadata ====================
    ws3 = wb.create_sheet("Document Metadata")
    
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 50
    
    ws3.merge_cells('A1:B1')
    ws3['A1'] = "DOCUMENT METADATA"
    ws3['A1'].font = header_font
    ws3['A1'].fill = header_fill
    ws3['A1'].alignment = center_alignment
    
    metadata = [
        ("AI System Name", ""),
        ("AI System Version", ""),
        ("Document Version", "1.0"),
        ("Date Created", ""),
        ("Last Updated", ""),
        ("Author", ""),
        ("Reviewer", ""),
        ("Approval Date", ""),
        ("Approved By", ""),
        ("Next Review Date", ""),
        ("Risk Classification", "☐ High-Risk (Annex III)"),
        ("Conformity Assessment Status", ""),
        ("EU Database Registration ID", ""),
    ]
    
    for row, (label, value) in enumerate(metadata, 3):
        ws3.cell(row=row, column=1, value=label).font = section_font
        ws3.cell(row=row, column=1).border = thin_border
        ws3.cell(row=row, column=1).fill = section_fill
        ws3.cell(row=row, column=2, value=value).border = thin_border
    
    # Save workbook
    output_path = "/Users/miachen/Library/CloudStorage/OneDrive-DTMasterCarbon/DT Master Mia Personal/5 Tech/AI act/EU_AI_Act_Article_13_Compliance_Template.xlsx"
    wb.save(output_path)
    print(f"Template created successfully: {output_path}")
    return output_path

if __name__ == "__main__":
    create_article13_template()

#!/usr/bin/env python3
"""
Create EU AI Act Annex XII Compliance Template
Transparency information for downstream providers integrating GPAI models
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_annex12_template():
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    subheader_font = Font(bold=True, size=11, color="FFFFFF")
    section_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")  # Green theme for GPAI
    subheader_fill = PatternFill(start_color="43A047", end_color="43A047", fill_type="solid")
    section_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    alt_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # ==================== Sheet 1: GPAI Model Documentation ====================
    ws1 = wb.create_sheet("GPAI Model Documentation")
    
    # Set column widths
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 35
    ws1.column_dimensions['C'].width = 50
    ws1.column_dimensions['D'].width = 40
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 30
    
    # Title
    ws1.merge_cells('A1:F1')
    ws1['A1'] = "EU AI ACT - ANNEX XII: TRANSPARENCY INFORMATION FOR DOWNSTREAM PROVIDERS"
    ws1['A1'].font = header_font
    ws1['A1'].fill = header_fill
    ws1['A1'].alignment = center_alignment
    ws1.row_dimensions[1].height = 30
    
    # Subtitle
    ws1.merge_cells('A2:F2')
    ws1['A2'] = "Technical Documentation for Providers of General-Purpose AI Models (Article 53(1)(b))"
    ws1['A2'].font = subheader_font
    ws1['A2'].fill = subheader_fill
    ws1['A2'].alignment = center_alignment
    
    # Headers
    headers = ["Ref.", "Requirement", "Description/Guidance", "Your Input", "Status", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Annex XII Requirements Data
    requirements = [
        # Section 1: General Description
        ("Section", "1. GENERAL DESCRIPTION OF THE GPAI MODEL - Annex XII(1)", "", "", "", ""),
        
        ("XII.1.a", "Intended Tasks", "Tasks that the model is intended to perform", "", "☐ Complete", ""),
        ("XII.1.a", "AI System Integration Types", "Type and nature of AI systems into which the model can be integrated", "", "☐ Complete", ""),
        ("XII.1.a", "Example Use Cases", "Illustrative examples of how the model can be used", "", "☐ Complete", "Recommended"),
        
        ("XII.1.b", "Acceptable Use Policy", "Policies defining acceptable and prohibited uses of the model", "", "☐ Complete", ""),
        ("XII.1.b", "Prohibited Uses", "Explicitly prohibited applications or use cases", "", "☐ Complete", ""),
        ("XII.1.b", "Use Restrictions", "Any restrictions on deployment contexts", "", "☐ Complete", ""),
        
        ("XII.1.c", "Release Date", "Date when the model was released", "", "☐ Complete", ""),
        ("XII.1.c", "Distribution Method", "How the model is distributed (API, download, etc.)", "", "☐ Complete", ""),
        ("XII.1.c", "Access Channels", "Platforms or channels where model is available", "", "☐ Complete", ""),
        
        ("XII.1.d", "Hardware Interaction", "How the model interacts with external hardware", "", "☐ Complete", "If applicable"),
        ("XII.1.d", "Software Interaction", "How the model interacts with external software", "", "☐ Complete", "If applicable"),
        ("XII.1.d", "API Specifications", "API endpoints and integration methods", "", "☐ Complete", ""),
        
        ("XII.1.e", "Software Versions", "Versions of relevant software for model usage", "", "☐ Complete", "If applicable"),
        ("XII.1.e", "Dependencies", "Required libraries, frameworks, and dependencies", "", "☐ Complete", ""),
        ("XII.1.e", "Compatibility Requirements", "Minimum system requirements", "", "☐ Complete", ""),
        
        ("XII.1.f", "Model Architecture", "Description of the model architecture (e.g., Transformer, CNN)", "", "☐ Complete", ""),
        ("XII.1.f", "Number of Parameters", "Total parameter count of the model", "", "☐ Complete", ""),
        ("XII.1.f", "Model Size", "Size of model files in storage", "", "☐ Complete", ""),
        
        ("XII.1.g", "Input Modality", "Types of inputs accepted (text, image, audio, etc.)", "", "☐ Complete", ""),
        ("XII.1.g", "Input Format", "Format specifications for inputs (JSON, etc.)", "", "☐ Complete", ""),
        ("XII.1.g", "Output Modality", "Types of outputs generated", "", "☐ Complete", ""),
        ("XII.1.g", "Output Format", "Format specifications for outputs", "", "☐ Complete", ""),
        
        ("XII.1.h", "Model License", "License under which the model is provided", "", "☐ Complete", ""),
        ("XII.1.h", "License Terms", "Key terms and conditions of the license", "", "☐ Complete", ""),
        ("XII.1.h", "Commercial Use Terms", "Terms for commercial deployment", "", "☐ Complete", ""),
        
        # Section 2: Model Elements and Development Process
        ("Section", "2. MODEL ELEMENTS AND DEVELOPMENT PROCESS - Annex XII(2)", "", "", "", ""),
        
        ("XII.2.a", "Integration Instructions", "Step-by-step instructions for integrating the model", "", "☐ Complete", ""),
        ("XII.2.a", "Required Infrastructure", "Hardware/cloud infrastructure requirements", "", "☐ Complete", ""),
        ("XII.2.a", "Technical Tools", "Tools and SDKs provided for integration", "", "☐ Complete", ""),
        ("XII.2.a", "Integration Code Examples", "Sample code for common integration scenarios", "", "☐ Complete", "Recommended"),
        
        ("XII.2.b", "Input Modality Details", "Detailed specification of input types", "", "☐ Complete", ""),
        ("XII.2.b", "Output Modality Details", "Detailed specification of output types", "", "☐ Complete", ""),
        ("XII.2.b", "Context Window Length", "Maximum context window/input size", "", "☐ Complete", ""),
        ("XII.2.b", "Maximum Output Size", "Maximum output length/size", "", "☐ Complete", ""),
        ("XII.2.b", "Token Limits", "Token limits for inputs and outputs", "", "☐ Complete", "If applicable"),
        ("XII.2.b", "Rate Limits", "API rate limits and quotas", "", "☐ Complete", "If applicable"),
        
        ("XII.2.c", "Training Data Type", "Types of data used for training (text, images, etc.)", "", "☐ Complete", "If applicable"),
        ("XII.2.c", "Training Data Provenance", "Sources of training data", "", "☐ Complete", "If applicable"),
        ("XII.2.c", "Data Curation Methodology", "Methods used to curate and process training data", "", "☐ Complete", "If applicable"),
        ("XII.2.c", "Testing Data Information", "Information about testing datasets", "", "☐ Complete", "If applicable"),
        ("XII.2.c", "Validation Data Information", "Information about validation datasets", "", "☐ Complete", "If applicable"),
        
        # Additional Recommended Information
        ("Section", "ADDITIONAL RECOMMENDED INFORMATION (Best Practices)", "", "", "", ""),
        
        ("REC.1", "Model Card", "Link to or inclusion of model card", "", "☐ Complete", "Recommended"),
        ("REC.2", "Performance Benchmarks", "Results on standard benchmarks", "", "☐ Complete", "Recommended"),
        ("REC.3", "Known Limitations", "Documented limitations of the model", "", "☐ Complete", "Recommended"),
        ("REC.4", "Bias Assessment", "Assessment of potential biases in model outputs", "", "☐ Complete", "Recommended"),
        ("REC.5", "Safety Evaluations", "Results of safety testing", "", "☐ Complete", "Recommended"),
        ("REC.6", "Red Team Testing", "Summary of adversarial testing performed", "", "☐ Complete", "Recommended"),
        ("REC.7", "Version History", "Changelog and version history", "", "☐ Complete", "Recommended"),
        ("REC.8", "Support Channels", "Technical support contact information", "", "☐ Complete", "Recommended"),
        ("REC.9", "Update Policy", "Policy for model updates and deprecation", "", "☐ Complete", "Recommended"),
    ]
    
    row = 5
    for req in requirements:
        if req[0] == "Section":
            ws1.merge_cells(f'A{row}:F{row}')
            ws1.cell(row=row, column=1, value=req[1])
            ws1.cell(row=row, column=1).font = section_font
            ws1.cell(row=row, column=1).fill = section_fill
            ws1.cell(row=row, column=1).alignment = Alignment(vertical='center')
            for col in range(1, 7):
                ws1.cell(row=row, column=col).border = thin_border
        else:
            for col, value in enumerate(req, 1):
                cell = ws1.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.alignment = wrap_alignment
                cell.border = thin_border
                if row % 2 == 0:
                    cell.fill = alt_fill
        row += 1
    
    # ==================== Sheet 2: Compliance Checklist ====================
    ws2 = wb.create_sheet("Compliance Checklist")
    
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 30
    
    ws2.merge_cells('A1:E1')
    ws2['A1'] = "ANNEX XII COMPLIANCE CHECKLIST"
    ws2['A1'].font = header_font
    ws2['A1'].fill = header_fill
    ws2['A1'].alignment = center_alignment
    
    checklist_headers = ["Annex Ref.", "Requirement", "Compliant?", "Evidence Location", "Responsible Person"]
    for col, header in enumerate(checklist_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    checklist_items = [
        ("XII(1)(a)", "Intended tasks documented", "☐ Yes / ☐ No", "", ""),
        ("XII(1)(a)", "AI system integration types specified", "☐ Yes / ☐ No", "", ""),
        ("XII(1)(b)", "Acceptable use policies defined", "☐ Yes / ☐ No", "", ""),
        ("XII(1)(c)", "Release date documented", "☐ Yes / ☐ No", "", ""),
        ("XII(1)(c)", "Distribution methods specified", "☐ Yes / ☐ No", "", ""),
        ("XII(1)(d)", "Hardware/software interaction described", "☐ Yes / ☐ No / ☐ N/A", "", ""),
        ("XII(1)(e)", "Relevant software versions listed", "☐ Yes / ☐ No / ☐ N/A", "", ""),
        ("XII(1)(f)", "Architecture described", "☐ Yes / ☐ No", "", ""),
        ("XII(1)(f)", "Number of parameters specified", "☐ Yes / ☐ No", "", ""),
        ("XII(1)(g)", "Input modality and format documented", "☐ Yes / ☐ No", "", ""),
        ("XII(1)(g)", "Output modality and format documented", "☐ Yes / ☐ No", "", ""),
        ("XII(1)(h)", "License specified", "☐ Yes / ☐ No", "", ""),
        ("XII(2)(a)", "Integration instructions provided", "☐ Yes / ☐ No", "", ""),
        ("XII(2)(a)", "Required infrastructure documented", "☐ Yes / ☐ No", "", ""),
        ("XII(2)(a)", "Technical tools listed", "☐ Yes / ☐ No", "", ""),
        ("XII(2)(b)", "Input/output modality detailed", "☐ Yes / ☐ No", "", ""),
        ("XII(2)(b)", "Maximum sizes (context window) specified", "☐ Yes / ☐ No", "", ""),
        ("XII(2)(c)", "Training data type documented", "☐ Yes / ☐ No / ☐ N/A", "", ""),
        ("XII(2)(c)", "Training data provenance documented", "☐ Yes / ☐ No / ☐ N/A", "", ""),
        ("XII(2)(c)", "Data curation methodology described", "☐ Yes / ☐ No / ☐ N/A", "", ""),
        ("Art. 53(1)(b)", "Documentation enables good understanding", "☐ Yes / ☐ No", "", ""),
        ("Art. 53(1)(b)", "Documentation enables compliance", "☐ Yes / ☐ No", "", ""),
    ]
    
    for row, item in enumerate(checklist_items, 4):
        for col, value in enumerate(item, 1):
            cell = ws2.cell(row=row, column=col, value=value)
            cell.font = normal_font
            cell.alignment = wrap_alignment
            cell.border = thin_border
            if row % 2 == 0:
                cell.fill = alt_fill
    
    # ==================== Sheet 3: Downstream Provider Info ====================
    ws3 = wb.create_sheet("Downstream Provider Info")
    
    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 50
    
    ws3.merge_cells('A1:B1')
    ws3['A1'] = "DOWNSTREAM PROVIDER INFORMATION"
    ws3['A1'].font = header_font
    ws3['A1'].fill = header_fill
    ws3['A1'].alignment = center_alignment
    
    ws3.merge_cells('A2:B2')
    ws3['A2'] = "Information about the downstream provider receiving this documentation"
    ws3['A2'].font = Font(italic=True, size=10)
    ws3['A2'].alignment = center_alignment
    
    downstream_info = [
        ("Downstream Provider Name", ""),
        ("Provider Registration Number", ""),
        ("Contact Person", ""),
        ("Contact Email", ""),
        ("Contact Phone", ""),
        ("Address", ""),
        ("Intended AI System", "Description of the AI system being developed"),
        ("Intended Use Case", "How the GPAI model will be used"),
        ("Risk Classification", "☐ High-Risk / ☐ Limited Risk / ☐ Minimal Risk"),
        ("Integration Date", ""),
        ("Documentation Version Received", ""),
        ("Date Documentation Received", ""),
    ]
    
    for row, (label, value) in enumerate(downstream_info, 4):
        ws3.cell(row=row, column=1, value=label).font = section_font
        ws3.cell(row=row, column=1).border = thin_border
        ws3.cell(row=row, column=1).fill = section_fill
        ws3.cell(row=row, column=2, value=value).border = thin_border
    
    # ==================== Sheet 4: Metadata ====================
    ws4 = wb.create_sheet("Document Metadata")
    
    ws4.column_dimensions['A'].width = 35
    ws4.column_dimensions['B'].width = 50
    
    ws4.merge_cells('A1:B1')
    ws4['A1'] = "GPAI MODEL PROVIDER - DOCUMENT METADATA"
    ws4['A1'].font = header_font
    ws4['A1'].fill = header_fill
    ws4['A1'].alignment = center_alignment
    
    metadata = [
        ("GPAI Model Name", ""),
        ("GPAI Model Version", ""),
        ("Provider Name", ""),
        ("Provider Address", ""),
        ("Provider Contact Email", ""),
        ("Authorised Representative (if outside EU)", ""),
        ("Document Version", "1.0"),
        ("Date Created", ""),
        ("Last Updated", ""),
        ("Author", ""),
        ("Reviewer", ""),
        ("Approval Date", ""),
        ("Approved By", ""),
        ("Next Review Date", ""),
        ("Systemic Risk Classification", "☐ Yes (Art. 51) / ☐ No"),
        ("Open Source Exception Applies", "☐ Yes (Art. 53(2)) / ☐ No"),
        ("Copyright Policy Reference", "Per Art. 53(1)(c)"),
        ("Training Data Summary URL", "Per Art. 53(1)(d)"),
    ]
    
    for row, (label, value) in enumerate(metadata, 3):
        ws4.cell(row=row, column=1, value=label).font = section_font
        ws4.cell(row=row, column=1).border = thin_border
        ws4.cell(row=row, column=1).fill = section_fill
        ws4.cell(row=row, column=2, value=value).border = thin_border
    
    # Save workbook
    output_path = "/Users/miachen/Library/CloudStorage/OneDrive-DTMasterCarbon/DT Master Mia Personal/5 Tech/AI act/EU_AI_Act_Annex_XII_Downstream_Provider_Template.xlsx"
    wb.save(output_path)
    print(f"Template created successfully: {output_path}")
    return output_path

if __name__ == "__main__":
    create_annex12_template()
